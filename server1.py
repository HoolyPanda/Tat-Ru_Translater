from pymystem3 import Mystem as stem
import openpyxl
import linguoData



class Server:
    def __init__(self):
        self.lang = "Ru"
        pass
    
    def __convertLangPart(self, lp):
        if "ADV" in lp:
            return "нар"
        elif "SPO" in lp:
            return "мест"
        elif "NUM" in lp:
            return "чис"
        elif "APRO" in lp:
            return "пм"
        elif "PR" in lp:
            return "пред"
        elif "V" in lp:
            return "гл"
        elif "S" in lp:
            return "сущ"
        elif "A" in lp:
            return "прил"
        else:
            return None
    
    def checkDictionaries(self, word, wordType = None):
        for dictionary in linguoData.dicts:
            if word in dictionary:
                self.passWord = True
                return dictionary[word]
            else:
                pass
        if word in linguoData.mestPrefx:
            return linguoData.mestPrefx[word]
        self.book = openpyxl.load_workbook(filename = "./tato-wordlist.xlsx")
        self.dict = self.book.get_sheet_by_name(self.book.get_sheet_names()[0])
        rows = self.dict.max_row
        for wordIndex in range(1, rows + 1):
            if self.dict.cell(wordIndex, 1).value == word:
                langPart = self.__convertLangPart(lp = wordType)
                if langPart is not None:
                    if self.dict.cell(wordIndex, 2).value != langPart:
                        self.passWord = False
                        return (linguoData.chastrechiPrist[langPart] + self.dict.cell(wordIndex,3))
                    else:
                        self.passWord = False
                        return (self.dict.cell(wordIndex, 3).value)
                else:
                    print("Error! " + wordType)
                pass
        print("No перевода для слова " + word + '\n' + "Добавить? (д/н)")
        a = input("Введите команду: ")
        if  a == "д":
            translation = input("      Введите перевода для слова " + word + ": ")
            wType = input("      Введите тип слова: ")
            self.dict.cell(rows + 1, 1).value = word
            self.dict.cell(rows + 1, 2).value = wType
            self.dict.cell(rows + 1, 3).value = translation
            self.dict = self.book.get_sheet_by_name(self.book.get_sheet_names()[1])
            self.dict.cell(rows + 1, 1).value = word
            self.dict.cell(rows + 1, 2).value = wType
            self.dict.cell(rows + 1, 3).value = translation
            self.book.save(filename = "./tato-wordlist.xlsx")
            self.book.close()
            self.passWord = False 
            return translation
        else:
            self.passWord = False
            return (word + "(NO TRANSLATION)")

    def Translate(self, text, lang = "Ru"):
        self.transaltedOutput = ""
        self.passWord = False
        if self.lang == "Ru":
            # a = stem().analyze(text)
            for buffWord in stem().analyze(text):
                if "analysis" in buffWord:
                    buff = buffWord['analysis'][0]
                    gr = buff["gr"]
                    lex = buff["lex"]
                    self.transaltedOutput += self.checkDictionaries(word = lex, wordType = gr)
                else:
                    if not self.passWord:
                        self.transaltedOutput += buffWord["text"]
                    else:
                        self.passWord = True
                    pass
            print(self.transaltedOutput)
        pass

