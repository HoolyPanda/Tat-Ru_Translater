# from pandas import ExcelFile
# from pandas import ExcelWriter
import pandas
from pymystem3 import Mystem as stem
# import json
# from xlrd import open_workbook
# import xlrd 
import openpyxl


class Server:
    def __init__(self):
        self.addTranslations = []
        pass

    def __TransalteComplexWord__(self, word:str):
        # fror debugging : word is братоубийца => титао-сай-ливэй
        if (self.lang == 'Ru'):
            s=stem()
            a = s.lemmatize(word)
            buff = ""
            i = 0
            for letter in word:
                buff += letter
                i += 1
                firstRoot = self.__translate__(s.lemmatize(buff)[0])
                if (i < len(word)):
                    if ((firstRoot is not None) & (word[i] == 'о')):
                        _word = word[(i+1):(len(word))]
                        buff = ""
                        i = 0
                        secRoot = self.__translate__(s.lemmatize(_word)[0])
                        for letter in _word:
                            buff += letter
                            i += 1
                            secRoot = self.__translate__(s.lemmatize(buff)[0])
                            a = s.analyze(buff)
                            if ((secRoot is not None)):
                                print(firstRoot + "-" + secRoot)
                            elif (i == len(_word)):
                                print("Нет перевода для слова: " + _word)
                                self.addTranslations.append(_word)
                                return True
                else:
                    print("Нет перевода для слова: " + word)
                    self.addTranslations.append(word)
                    return False

            pass
        elif (self.lang == 'Ta'):
            pass
        pass

    
    def __CompleteToNoun(self, word:str):
        return "сай-" + word
    def __CompleteToAdj(self, word:str):
        return "Пара-" + word
    def __CompleteToVerb(self, word:str):
        return "гай-" + word
    def __CompleteToAdv(self, word:str):
        return "зай-" + word

    def __checkIfComplexWord(self, word:str):
        pass

    def __convertLangPart(self, lp):
        if "ADV" in lp:
            return "нар"
        elif "SPO" in lp:
            return "мест"
        elif "NUM" in lp:
            return "чис"
        elif "APRO" in lp:
            return "пм"
        elif "PR" in lp:
            return "пред"
        elif "V" in lp:
            return "гл"
        elif "S" in lp:
            return "сущ"
        elif "A" in lp:
            return "прил"
        else:
            return "-"

    def __TranslateSentance__ (self,sent:str):
        
        out = ""
        if (self.lang == "Ru"):
            word = ""
            a = stem().analyze(sent)
            for sword in stem().analyze(sent):
                if ((sword.get("text") != ".") & (sword.get("text") != " ") & (sword.get("text") != "\n")):
                    word = sword.get("analysis")[0]
                    langPart = sword.get("analysis")[0].get("gr")[0]
                    o = self.__translate__(word)
                    if (o):
                        out += (o + "-")
                elif ((sword.get("text") != " ") & (sword.get("text") != "\n")):
                    out = out[0:(len(out) - 1)]
                    out += sword.get("text") 
                elif (sword.get("text") == "\n"):
                    out = out[:-1]
                    out += "\n"
                pass
            if (len(self.addTranslations) != 0):
                for aWord in self.addTranslations:
                    print("Ctrl + C to quit")
                    print(aWord)
                    wordType = input("      Введите тип слова: ") 
                    wortdTranslate = input("      Введите перевод слова: ") 
                    wb = openpyxl.load_workbook(filename = "./tato-wordlist.xlsx")
                    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[1])
                    rows = sheet.max_row + 1
                    sheet.cell(rows, 1).value = wortdTranslate
                    sheet.cell(rows, 2).value = wordType
                    sheet.cell(rows, 3).value = aWord
                    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
                    rows = sheet.max_row + 1
                    sheet.cell(rows, 1).value = aWord
                    sheet.cell(rows, 2).value = wordType
                    sheet.cell(rows, 3).value = wortdTranslate
                    wb.save(filename = "./tato-wordlist.xlsx")
                    wb.close()
            else:
                print ( sent + "\n" + out)
                pass
        elif (self.lang == "Ta"):
            pass

    def __translate__(self, word):
        for wordIndex in self.wl.index:
                a = self.wl["Слово"][wordIndex]
                if (self.wl["Слово"][wordIndex] == word.get('lex')):
                    lp = self.__convertLangPart(lp=word.get("gr"))
                    if self.wl["Часть речи"][wordIndex] == lp:
                        return (self.wl['Перевод'][wordIndex]) 
                    elif lp == "нар":
                        
                        pass
        unicWord = True
        for subWord in self.addTranslations:
            if subWord == word:
                unicWord = False
        if unicWord:
            self.addTranslations.append(word)
        return None


    def Translate(self, lang:str ,word:str):
        self.lang = lang
        if (lang == "Ta"):
            self.wl = pandas.read_excel("./tato-wordlist.xlsx" , sheet_name='Тато - Русский')
            self.wlIndex = self.wl.index
            if ('-' in word):
                if ('.' in word):
                    self.__TranslateSentance__(sent=word)
                else:
                    self.__TransalteComplexWord__(word=word[0])
                    pass
        elif (lang == 'Ru'):
            self.wl = pandas.read_excel("./tato-wordlist.xlsx" , sheet_name='Русский - Тато')
            s=stem()
            if (' ' in word):
                self.__TranslateSentance__(word)
            else:
                lemm = s.lemmatize(word)
                out = self.__translate__(word=lemm[0])
                if ( (out is None) & ( True) ):
                    self.__TransalteComplexWord__(word)
                else:
                    print( word + " => " + out)